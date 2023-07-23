import io
import json
import logging
import os
from datetime import datetime
from typing import Optional, Union

import boto3
import numpy as np
import openpyxl
import pandas as pd
from boto3.dynamodb.conditions import Key
from boto3.dynamodb.types import TypeDeserializer
from openpyxl.utils.dataframe import dataframe_to_rows

# ロギングの初期設定
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

s3 = boto3.client("s3")
dynamodb = boto3.resource("dynamodb")


# 曜日
WEEKDAY_JP: dict[int, str] = {
    0: "月",
    1: "火",
    2: "水",
    3: "木",
    4: "金",
    5: "土",
    6: "日",
}

# 勤務形態
WORK_CODE: dict[str, str] = {
    "01": "client_onsite",
    "02": "client_offsite",
    "10": "in-house_onsite",
    "11": "in-house_offsite",
}


# カスタムエラーを定義
class WorkforceBuddyException(Exception):
    pass


def lambda_handler(event: dict, context: dict) -> dict:
    """
    Lambda関数ハンドラ

    Args:
        event (dict)
        context (dict)

    Returns:
        dict: レスポンス
    """
    try:
        logger.info(f"event: {event}")
        res: dict = logic(event)

    except WorkforceBuddyException:
        raise WorkforceBuddyException

    except Exception as err:
        logger.error(f"想定外のエラーが発生しました\n{err}")
        raise WorkforceBuddyException

    return res


def logic(event: dict) -> dict:
    """
    メインロジック

    Args:
        event (dict)

    Returns:
        dict: レスポンス
    """
    # ファイル情報の読み出し
    deserializer = TypeDeserializer()
    try:
        user_config: dict = {
            k: deserializer.deserialize(v)
            for k, v in event["user_config"]["Item"].items()
        }
        template_config: dict = {
            k: deserializer.deserialize(v)
            for k, v in event["template_config"]["Item"].items()
        }
        work_info: dict = event["work_info"]
        work_month: str = event["work_months"]
        bucket_name: str = os.environ["BUCKET_NAME"]
        table_name: str = os.environ["TABLE_NAME"]
    except Exception as err:
        logger.error("環境情報の読み出しに失敗しました\n{err}")
        raise WorkforceBuddyException

    # 勤務データをDBから取得
    table = dynamodb.Table(table_name)
    work_data: dict = table.query(
        KeyConditionExpression=Key("id").eq(work_info["user_id"])
        & Key("SK").begins_with(f"WorkData#{work_month}"),
    )["Items"]
    work_df: pd.DataFrame = pd.DataFrame(work_data)

    # 勤務データを必要な形式に加工
    converted_work_df: pd.DataFrame = convert_work_data(
        work_month, work_df, user_config
    )

    # テンプレートファイルの読み込み
    template_path: str = f"template/{template_config['name']}"
    template_file: Optional[bytes] = None
    try:
        template_file = (
            s3.get_object(Bucket=bucket_name, Key=template_path)
            .get("Body")
            .read()
        )
    except Exception as err:
        logger.error(f"テンプレートファイルの取得に失敗しました\n{err}")
        raise WorkforceBuddyException

    # テンプレートファイルを取得できなかった場合
    if not template_file:
        logger.error("テンプレートファイルの取得に失敗しました")
        raise WorkforceBuddyException

    # 勤務表の生成
    work_schedule_file: bytes = create_work_schedule(
        template_file,
        template_config,
        user_config,
        work_month,
        converted_work_df,
    )

    # 勤務表をアップロード
    work_schedule_object_name: str = (
        f"{user_config['id']}_{'_'.join(work_month.split('-'))}.xlsx"
    )
    work_schedule_path = f"work_schedule/{work_schedule_object_name}"
    try:
        s3.put_object(
            Bucket=bucket_name, Body=work_schedule_file, Key=work_schedule_path
        )
    except Exception as err:
        logger.error(f"ファイルのアップロードに失敗しました\n{err}")
        raise WorkforceBuddyException

    # レスポンスを生成
    response: dict = create_response(
        work_month, bucket_name, work_schedule_object_name
    )

    return response


def get_work_data(table_name: str, user_id: str, work_month: str) -> dict:
    """
    勤務データをDBから取得

    Args:
        table_name (str): テーブル名
        user_id (str): 社員番号
        work_month (str): 勤務月(ex: '2023-07')

    Returns:
        dict: 勤務データ
    """
    table = dynamodb.Table(table_name)

    # 勤務データの取得
    work_data: dict = table.query(
        KeyConditionExpression=Key("id").eq(user_id)
        & Key("SK").begins_with(f"WorkData#{work_month}"),
    )["Items"]

    return work_data


def convert_work_data(
    work_month: str, df: pd.DataFrame, user_config: dict
) -> pd.DataFrame:
    """
    勤務データを必要な形式に加工

    Args:
        work_month (str): 勤務月(ex: '2023-07')
        df (pd.DataFrame): 勤務データ
        user_config (dict): ユーザ設定

    Returns:
        pd.DataFrame: 加工済み勤務データ
    """
    # 勤務表に記載する勤務形態を選別
    df = df.loc[df["work_code"].isin(["01", "02"]), :]

    # 勤務日の型変換(datetime(yyyy, mm, dd, 0, 0))
    df["datetime"] = df["datetime"].map(
        lambda x: datetime.fromisoformat(x), na_action="ignore"
    )

    # 開始時刻の型変換(str->datetime(yyyy, mm, dd, hh, MM))
    df["start_datetime"] = df["start_datetime"].map(
        lambda x: datetime.fromisoformat(x), na_action="ignore"
    )

    # 終了時刻の型変換(str->datetime(yyyy, mm, dd, hh, MM))
    df["end_datetime"] = df["end_datetime"].map(
        lambda x: datetime.fromisoformat(x), na_action="ignore"
    )

    # 開始時刻の生成(timedelta(xxdays, xxhours, xxminutes))
    df["start_timedelta"] = df["start_datetime"] - df["datetime"]

    # 終了時刻の生成(timedelta(xxdays, xxhours, xxminutes))
    df["end_timedelta"] = df["end_datetime"] - df["datetime"]

    # 開始時間の生成(timedelta(xxhours, xxminutes))
    time_sharing = int(user_config["time_sharing"])
    df["start_time"] = (df["start_datetime"] - df["datetime"]).map(
        lambda x: print_timedelta(x, time_sharing),
        na_action="ignore",
    )

    # 終了時間の生成(timedelta(xxhours, xxminutes))
    df["end_time"] = (df["end_datetime"] - df["datetime"]).map(
        lambda x: print_timedelta(x, time_sharing),
        na_action="ignore",
    )

    # 欠損した日付を補完
    one_month_df = create_one_month_dataframe(work_month)
    df = pd.merge(one_month_df, df, on="datetime", how="left")

    # 勤務日の生成
    df["work_day"] = df["datetime"].map(lambda x: str(x.day))

    # 勤務曜日の生成
    df["work_weekday"] = df["datetime"].map(lambda x: WEEKDAY_JP[x.weekday()])

    # NaNをNoneに変換
    df = df.replace({pd.NA: None})

    return df


def print_timedelta(
    time: pd.Timedelta, time_sharing: int
) -> Union[str, float]:
    """
    timedelta形式のデータを文字列型に変換する

    Args:
        time (pd.Timedelta): 時刻
        time_sharing (int): 時刻まるめ(ex: 15 -> 15分単位で切り捨て)

    Returns:
        Union[str, float]: hh:mm形式の時刻文字列
    """
    # naの場合nanを返す
    if time is np.nan:
        return np.nan

    # timedeltaで渡された時間の時分を算出
    total_seconds = int(time.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds - (hours * 3600)) // 60

    # 時刻をまるめる
    sharing_minutes = (minutes // time_sharing) * time_sharing

    return f"{hours:>02}:{sharing_minutes:>02}"


def create_one_month_dataframe(year_month: str) -> pd.DataFrame:
    """
    指定したひと月の範囲の日にちを持ったDataFrameを生成する

    Args:
        year_month (str): 特定の年月(ex: '2023-05')

    Returns:
        pd.DataFrame: ひと月の範囲の日にちを持ったDataFrame
    """
    # 開始日と終了日を取得
    start_date: datetime = datetime.strptime(year_month, "%Y-%m")
    end_date: datetime = start_date + pd.DateOffset(months=1, days=-1)

    # 日付の範囲を生成
    date_range: pd.date_range = pd.date_range(
        start=start_date, end=end_date, freq="D"
    )

    # データフレームを作成
    df: pd.DataFrame = pd.DataFrame({"datetime": date_range})

    return df


def create_work_schedule(
    template_file: bytes,
    template_config: dict,
    user_config: dict,
    work_month: str,
    df: pd.DataFrame,
) -> bytes:
    """
    テンプレートファイルから勤務表を作成

    Args:
        template_file (bytes): テンプレートファイル
        template_config (dict): 作成する勤務表の設定
        user_config (dict): ユーザ設定
        work_month (str): 勤務月(ex: '2023-07')
        df (pd.DataFrame): 勤務データ

    Returns:
        bytes: 勤務表
    """
    # テンプレート設定を読み込み
    year_month_formats: dict = json.loads(
        template_config["year_month_formats"]
    )
    year_month_cells: dict = json.loads(template_config["year_month_cells"])
    start_cells: dict = json.loads(template_config["start_cells"])
    user_name_cell: str = template_config["user_name_cell"]

    # 日付フォーマットの変換
    year_month: datetime = datetime.strptime(work_month, "%Y-%m")
    year: str = str(year_month.year)
    month: str = str(year_month.month)
    year_months: dict = {}
    for k, v in year_month_formats.items():
        year_months[k] = v.format(year=year, month=month)

    # テンプレートファイルの読み込み
    wb = openpyxl.load_workbook(io.BytesIO(template_file))
    ws = wb.worksheets[0]

    # 年月の書き込み
    for key, cell in year_month_cells.items():
        ws[cell] = year_months[key]

    # 氏名の書き込み
    ws[user_name_cell] = user_config.get("user_name")

    # 勤務データの書き込み
    for column, cell in start_cells.items():
        for i, value in enumerate(
            dataframe_to_rows(df[[column]], index=False, header=False)
        ):
            current_cell = ws[cell].offset(i)
            current_cell.value = value[0]

    # 勤務表の書き出し
    work_schedule_file: Optional[bytes] = None
    with io.BytesIO() as file:
        wb.save(file)
        work_schedule_file = file.getvalue()

    return work_schedule_file


def create_response(
    work_month: str, bucket_name: str, object_name: str
) -> dict:
    """
    関数の返却値を生成する

    Args:
        work_month (str): 勤務月(ex: '2023-07')
        bucket_name (str): アップロード先S3バケット名
        object_name (str): アップロードしたファイル名

    Returns:
        dict:
            bucket_name (str): アップロード先S3バケット名
            object_name (str): アップロードしたファイル名
    """
    res: dict = {
        "work_month": work_month,
        "bucket_name": bucket_name,
        "object_name": object_name,
    }

    return res
